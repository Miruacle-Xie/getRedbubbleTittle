# -*- coding: utf-8 -*-
"""
Created on Mon Jul 25 09:09:52 2022

@author: Administrator
"""
import pandas as pd
import nltk
from nltk import word_tokenize
from nltk.util import ngrams
import os
import re
from openpyxl import load_workbook
import datetime
import sys


gl_WORDFREQFILE = "主题词频"
gl_WORDFREQFILEPATH = ""
gl_FILTERWORD = "过滤词"
gl_FILTERWORDPATH = ""
gl_LIB = "Lib"
gl_LIBPATH = ""
gl_OUTPUT = "output"
gl_OUTPUTPATH = ""


def openExcel(fileName):
    fileName = "redbubble.xlsx"
    return pd.read_excel(fileName.replace("\"", ""), sheet_name=0)


def readTittle(df, column=3):
    return df.iloc[:, column]


def readFilterWord():
    df = pd.read_excel(os.path.normpath(gl_FILTERWORDPATH))
    return df.iloc[:, 0]


def delFilterWord(text, filterWord):
    for word in filterWord:
        # print(word)
        # print(":\n")
        if word == '|':
            wordtmp = "\|"
        else:
            wordtmp = word
        word = '( '+wordtmp+'(s|es)? )|( '+wordtmp + \
            '(s|es)?$)|(^'+wordtmp+'(s|es)? )'
        repl = " "
        # print(word)
# =============================================================================
#         if word == '|':
#             word = '( \| )|( \|$)|(^\| )'
#             repl = " "
#         else:
#             word = '\\b' + word + '(s|es)?' + "\\b"
#             repl = ""
# =============================================================================
        text = text.str.replace(word, repl, case=False, regex=True)
        # print(text)
    return text.str.replace(" {2,}", " ", case=False, regex=True).str.strip()


def createNgrams(text, number):
    textNgrams = []
    for line in text:
        token = word_tokenize(line)
        tmpText = [" ".join(x) for x in list(ngrams(token, number))]
        word = [w for w in tmpText if not re.findall(
            '(^[^\w\s]|[^\w\s]$)', w)]  # 仅保留单词数字
        textNgrams = textNgrams + word

    return textNgrams


def extractSubject(fileName):
    # print(today.strftime("%y-%m-%d"))
    # sheetName = pd.read_excel(fileName.replace("\"", ""), sheet_name=None)
    originalData = pd.read_excel(fileName.replace("\"", ""))
    # print(originalData)
    # print('----------------------\n')
    tittle = readTittle(originalData)
    # print(tittle)
    FilterWord = readFilterWord().tolist()
    FilterWord.sort(key=lambda x: x.count(' '), reverse=True)
    # print(FilterWord)
    subjectName = delFilterWord(tittle, FilterWord)
    subjectNameSplit = subjectName.str.split(",|/", expand=True)
    for i in range(len(subjectNameSplit.columns)):
        subjectNameSplit[i] = subjectNameSplit[i].str.strip()
    # excelWriter = createExcelWriter(fileName)
    excelWriter = createExcelWriter(gl_OUTPUTPATH+fileName)
    tittle = tittle.rename('原标题')
    tittle.to_excel(excelWriter, index=False)
    subjectNameSplit.to_excel(excelWriter, startcol=1, index=False)
    savecloseExcelWrite(excelWriter)
    excelWriter = createExcelWriter(gl_WORDFREQFILEPATH)
    wordFreqReport(excelWriter, subjectName, fileName, 2, 0)
    wordFreqReport(excelWriter, subjectName, fileName, 1, 3)
    savecloseExcelWrite(excelWriter)


def wordFreqReport(excelWriter, textValue, fileName, ngram, col):
    subjectFreq = wordFreq(textValue, ngram)
    df = pd.DataFrame(subjectFreq.items(), columns=[
                      'word', 'freq']).sort_values(by='freq', ascending=False)
    df.to_excel(excelWriter, sheet_name=os.path.splitext(
        fileName)[0][:32], index=False, startcol=col)


def createExcelWriter(fileName):
    if os.path.isfile(fileName):
        if os.path.getsize(fileName):
            book = load_workbook(fileName)
            excelWriter = pd.ExcelWriter(
                fileName, engine='openpyxl')
            excelWriter.book = book
        else:
            excelWriter = pd.ExcelWriter(
                fileName, engine='openpyxl')
    else:
        excelWriter = pd.ExcelWriter(fileName, engine='openpyxl')
    return excelWriter


def savecloseExcelWrite(excelWriter):
    excelWriter.save()
    excelWriter.close()


def excelWrite(fileName, textValue, sheetName='', mode='a'):
    if mode == 'a':
        if os.path.isfile(fileName):
            if os.path.getsize(fileName):
                book = load_workbook(fileName)
                excelWriter = pd.ExcelWriter(fileName, engine='openpyxl')
                excelWriter.book = book
            else:
                excelWriter = pd.ExcelWriter(fileName, engine='openpyxl')
        else:
            excelWriter = pd.ExcelWriter(fileName, engine='openpyxl')
    else:
        excelWriter = pd.ExcelWriter(fileName, engine='openpyxl')
    if sheetName == '':
        textValue.to_excel(excelWriter, index=False)
    else:
        textValue.to_excel(excelWriter, sheet_name=sheetName, index=False)
    excelWriter.save()
    excelWriter.close()


def wordFreq(originalData, ngrams):
    wordgram = createNgrams(originalData, ngrams)
    fdist = nltk.FreqDist(wordgram)
    # print(fdist.items())
    # print(len(fdist.items()))
    # print(fdist.keys())
    # print(fdist.max())
    # print(fdist[fdist.max()])
    return fdist


def main():
    dirpath = "F:\\JetBrains\\redbubbleProcess\\标题数据"
    dirpath = input("\n需要的文件夹路径：\n")
    today = datetime.date.today().strftime("%y-%m-%d")
    global gl_WORDFREQFILEPATH
    global gl_FILTERWORDPATH
    global gl_LIBPATH
    global gl_OUTPUTPATH
    gl_OUTPUTPATH = dirpath + '\\' + gl_OUTPUT + '\\'
    if not os.path.isdir(gl_OUTPUTPATH):
        os.mkdir(gl_OUTPUTPATH)
    os.chdir(os.path.dirname(os.path.realpath(__file__)))
    os.chdir(os.path.dirname(sys.executable))
    gl_LIBPATH = os.getcwd()+'\\' + gl_LIB + '\\'
    gl_WORDFREQFILEPATH = gl_LIBPATH+gl_WORDFREQFILE+today+'.xlsx'
    gl_FILTERWORDPATH = gl_LIBPATH+gl_FILTERWORD+'.xlsx'
    print(gl_WORDFREQFILEPATH)
    print(gl_FILTERWORDPATH)
    print(gl_OUTPUTPATH)
    # return
    # dirFiles = os.listdir(os.getcwd())
    # dirFiles = os.listdir(dirpath.replace("\"", ""))
    os.chdir(dirpath)
    dirFiles = os.listdir(re.sub("\"|\'", "", dirpath))
    # print(dirFiles)
    xlsxfile = [file for file in dirFiles if os.path.splitext(file)[1] in [
        '.xlsx']]
    # print(xlsxfile)
    for fileName in xlsxfile:
        if fileName != gl_FILTERWORD+'.xlsx' and fileName != gl_WORDFREQFILE+today+'.xlsx':
            print(fileName)
            extractSubject(fileName)
    input("运行结束,按回车键结束")


def test():
    df = openExcel(1)
    print("df:\n%s" % df)
    tittle = readTittle(df)
    textClear = delFilterWord(tittle, readFilterWord())
    # print(textClear)
    print("\ntextClear:\n%s" % textClear)
    # expand=True 将拆分出来的内容分别作为单独一列， 然后根据切片取所需那一列
    texttmp = textClear.str.split(",|/", expand=True)
    # print("\ntexttmp:\n%s"%texttmp)
    for i in range(len(texttmp.columns)):
        texttmp[i] = texttmp[i].str.strip()
    print(texttmp)
    blend = pd.concat([df, texttmp], axis=1)
    print("\ndf+texttmp:\n%s" % blend)
    path = os.getcwd()
    filterWordCorpus = path + "\\" + "filterWord1.xlsx"
    writer = pd.ExcelWriter(filterWordCorpus)
    # header = None：数据不含列名，index=False：不显示行索引（名字）
    texttmp.to_excel(writer, header=None, index=False)
    print(len(texttmp.columns))
    df.to_excel(writer, startcol=len(texttmp.columns), index=False)
    writer.save()
    # print("len(textClear):{}".format(len(textClear)))
    # bigram = createNgrams(textClear, 2)
    # print(bigram)
    # fdist = nltk.FreqDist(bigram)
    # # print(fdist.items())
    # print(len(fdist.items()))
    # print(fdist.keys())
    # print(fdist.max())
    # print(fdist[fdist.max()])


if __name__ == '__main__':
    # test()
    try:
        main()
    except Exception as e:
        input(str(e)+"\n\n运行异常,按回车键结束")
    
